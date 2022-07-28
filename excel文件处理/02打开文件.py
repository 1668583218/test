from openpyxl import load_workbook

# 打开表格
wb = load_workbook("./01.xlsx")
# 获取所有sheet页
ws_name = wb.sheetnames
# 获取第一页
ws = wb[ws_name[0]]
# # 获取单元格对象
# cell_1 = ws['A1']
# cell_2 = ws.cell(row=1, column=1)
# print(cell_1)
# print(cell_2)
#
# # 取得这个单元格的内容，只需要在结尾加上value属性
# value_1 = ws['A1'].value
# value_2 = ws.cell(row=1, column=1).value
# print(value_1)
# print(value_2)

# # 多个单元格的获取一般需要用到列表切片的知识或者使用for循环来进行
# # 访问A1至C3范围单元格
# cell_range = ws['A1':'C3']
# print(cell_range)
# # 访问A列所有存在数据的单元格
# colA = ws['A']
# # 访问A列到C列所有存在数据的单元格
# col_range = ws['A:C']
# # 访问第1行所有存在数据的单元格
# row1 = ws[1]
# # 访问第1行至第5行所有存在数据的单元格
# row_range = ws[1:5]
#
# # 注意，上述cell_range等对象都是<class 'tuple'>类型的。如果先获取这些单元格中的值，我们可以这样
# for each_cell in cell_range:
#     for each in each_cell:
#      print(each.value)
#
# for each_cell in colA:
#     print(each_cell.value)

# # 实战1
# item = []
# for row in ws.iter_rows(min_row=6, max_col=14, max_row=38):
#     x = 0
#     list1 = []
#     for cell in row:
#         list1.append(cell.value)
#     item.append(tuple(list1))
# print(item)
# 实战2
# 获取需要多少行数据
star = int(input("请输入开始行："))
end = int(input("请输入结束行："))

# 获取B列的数据
colB = ws['B{}:B{}'.format(star, end)]
listB = []
for i in colB:
    for x in i:
        listB.append(x.value)
print(listB)

# 获取C列的数据
colC = ws['C{}:C{}'.format(star, end)]
listC = []
for i in colC:
    for x in i:
        listC.append(x.value)
print(listC)

# 将B、C两列的数据添加到最终列表
item = []
for z in range(end - star):
    list1 = []
    list1.append(listB[z])
    list1.append(listC[z])
    item.append(tuple(list1))

print(item)