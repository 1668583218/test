from openpyxl import Workbook
# 创建一个 workbook
wb = Workbook()
# 获取被激活的 worksheet
ws = wb.active
# 设置单元格内容
ws['A1'] = 42
# 设置一行内容
ws.append([1, 2, 3])
# python 数据类型可以被自动转换
import datetime
ws['A3'] = datetime.datetime.now()
# 保存 Excel 文件
wb.save("sample.xlsx")